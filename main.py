import os
import cv2
import openpyxl
import easyocr
from ultralytics import YOLO
from datetime import datetime, timedelta
import multiprocessing as mp
import logging
import schedule
import time
import smtplib
import ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Set up logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    filename='vehicle_detection.log',
                    filemode='w')

logger = logging.getLogger(__name__)

# Config variables
SHOW_LIVE = True
PLATE_CONF_MIN = 0.65
VEHICLE_CONF_MIN = 0.5
VIDEO_SOURCE = "input_5.mkv"
RESIZE_FACTOR = 2
TIME_FORMAT = '%Y%m%d_%H%M%S.%f'
OUTPUT_DIR = "number_plates"
EXCEL_HEADERS = ["S. No.", "Object ID", "HSRP Detected", "Middle Conf.", "Middle Frame", "Second Last Conf.", "Second Last Frame", "Middle OCR", "Second Last OCR"]
VEHICLE_CLASSES = {2, 3, 5, 7}

# Usage
sender_email = "your_email@dccmail.in"
to_emails = ["recipient1@example.com", "recipient2@example.com"]
cc_emails = ["cc_recipient1@example.com", "cc_recipient2@example.com"]
password = "your_password"
subject = "Email with Attachment"
body = "Please find the attached file."
host = "smtp.gmail.com"
port = 465

# Initialize models
try:
    vehicle_model = YOLO('yolo11n.pt')
    plate_model = YOLO('best28.pt')
except Exception as e:
    logger.error(f"Error loading YOLO models: {e}")
    raise

logger.info("Models loaded successfully")

# Initialize EasyOCR reader
try:
    reader = easyocr.Reader(['en'], gpu=False)
except Exception as e:
    logger.error(f"Error initializing EasyOCR: {e}")
    raise


def ensure_dir(directory):
    os.makedirs(directory, exist_ok=True)
    logger.debug(f"Ensured directory exists: {directory}")


def get_date_folder():
    return datetime.now().strftime('%Y-%m-%d')


def get_output_dirs():
    date_folder = get_date_folder()
    output_dir = os.path.join(OUTPUT_DIR, date_folder)
    plates_dir = os.path.join(output_dir, "plates")
    frames_dir = os.path.join(output_dir, "frames")
    return output_dir, plates_dir, frames_dir


def recognize_plate(plate_img):
    try:
        ocr_result = reader.readtext(plate_img)
        if ocr_result:
            return ocr_result[0][1], ocr_result[0][2]  # text and confidence
        return None, None
    except Exception as e:
        logger.error(f"Error in plate recognition: {e}")
        return None, None


def save_image(directory, filename, image):
    ensure_dir(directory)
    cv2.imwrite(os.path.join(directory, filename), image)
    logger.debug(f"Saved image: {os.path.join(directory, filename)}")


def get_plate(photo, obj_id):
    timestamp = datetime.now().strftime(TIME_FORMAT)
    _, plates_dir, _ = get_output_dirs()
    detections = plate_model.predict(source=photo, conf=PLATE_CONF_MIN, show=SHOW_LIVE)

    for detection in detections:
        if detection is not None:
            for xx, yx, xy, yy, detected_conf, _ in detection.boxes.data.tolist():
                plate = photo[int(yx):int(yy), int(xx):int(xy)]
                save_image(os.path.join(plates_dir, str(obj_id)), f"{timestamp}.jpg", plate)
                logger.info(f"Plate detected for object {obj_id} at {timestamp}")
                return timestamp
    logger.debug(f"No plate detected for object {obj_id}")
    return timestamp


def process_frame(frame, result):
    timestamp = datetime.now().strftime(TIME_FORMAT)
    _, plates_dir, frames_dir = get_output_dirs()

    for obj in result.boxes.data.tolist():
        try:
            # Adapt the unpacking to match the actual structure (6 values instead of 7)
            x1, y1, x2, y2, obj_id, conf, obj_class = (map(float, obj))

            if obj_class in VEHICLE_CLASSES:
                vehicle = frame[int(y1):int(y2), int(x1):int(x2)]
                timestamp = get_plate(vehicle, obj_id)
                save_image(os.path.join(frames_dir, str(obj_id)), f"{timestamp}.jpg", vehicle)
                logger.info(f"Processed frame for object {obj_id} at {timestamp}")
        except ValueError as e:
            logger.warning(f"Skipping object due to unexpected data format: {obj}")
            continue


def vehicle_detection(frame_queue, result_queue):
    logger.info("Starting vehicle detection process")
    while True:
        for result in vehicle_model.track(source=VIDEO_SOURCE, conf=VEHICLE_CONF_MIN, stream=True, show=SHOW_LIVE):
            if result is not None:
                frame_queue.put((result.orig_img, result))
                if result_queue.qsize() > 0:
                    processed_result = result_queue.get()


def plate_detection(frame_queue, result_queue):
    logger.info("Starting plate detection process")
    while True:
        frame, result = frame_queue.get()
        if frame is None:
            break
        process_frame(frame, result)
        result_queue.put(result)


def save_blank_excel(output_file):
    workbook = openpyxl.Workbook()
    workbook.active.append(EXCEL_HEADERS)
    workbook.save(output_file)
    logger.info(f"Created blank Excel file: {output_file}")


def send_email_with_attachment(sender_email, to_emails, cc_emails, password, subject, body, filename):
    logger.info('Preparing to send email...')
    logger.info('Sender: %s', sender_email)
    logger.info('To: %s', ', '.join(to_emails))
    logger.info('Cc: %s', ', '.join(cc_emails))
    logger.info('Subject: %s', subject)
    logger.info('Attachment: %s', filename)

    # Create a multipart message
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = ", ".join(to_emails)
    message["Cc"] = ", ".join(cc_emails)
    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(body, "plain"))
    logger.debug('Email body attached')

    # Open the file in binary mode
    try:
        with open(filename, "rb") as attachment:
            # Add file as application/octet-stream
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        logger.debug('File %s read successfully', filename)
    except IOError as e:
        logger.error('Failed to read attachment file: %s', e)
        raise

    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)
    logger.debug('File encoded successfully')

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message
    message.attach(part)
    logger.debug('Attachment added to message')

    # Convert message to string
    text = message.as_string()

    # Combine all recipients
    all_recipients = to_emails + cc_emails

    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    try:
        with smtplib.SMTP(host, port) as server:
            logger.info('Connecting to SMTP server...')
            server.ehlo()  # Can be omitted
            server.starttls(context=context)
            server.ehlo()  # Can be omitted
            server.login(sender_email, password)
            logger.info('Logged in successfully')
            server.sendmail(sender_email, all_recipients, text)
            logger.info('Email sent successfully!')
    except smtplib.SMTPException as e:
        logger.error('An error occurred while sending the email: %s', e)
        raise

    logger.info('Email sending process completed')


def run_ocr_and_save_to_excel(date):
    logger.info(f"Starting OCR process for {date}")

    input_dir = os.path.join(OUTPUT_DIR, date)
    plates_dir = os.path.join(input_dir, "plates")
    output_file = os.path.join(input_dir, f"output_{date}.xlsx")

    if not os.path.exists(output_file):
        save_blank_excel(output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    for obj_id in os.listdir(plates_dir):
        obj_dir = os.path.join(plates_dir, obj_id)
        image_files = sorted(os.listdir(obj_dir))  # Sort filenames

        # Determine the middle and second to last images
        indices_to_process = []
        if len(image_files) > 1:
            middle_index = len(image_files) // 2
            second_last_index = len(image_files) - 2
            indices_to_process = [middle_index, second_last_index]

        results = []  # To hold results for the current object ID

        for index in indices_to_process:
            if index < len(image_files):
                image_file = image_files[index]
                image_path = os.path.join(obj_dir, image_file)
                plate_img = cv2.imread(image_path)

                text, confidence = recognize_plate(plate_img)
                results.append((text, confidence, image_file))  # Store results

        # If we have results, write them to the Excel sheet
        if results:
            text1, confidence1, image_file1 = results[0] if len(results) > 0 else (None, None, None)
            text2, confidence2, image_file2 = results[1] if len(results) > 1 else (None, None, None)

            row = [
                sheet.max_row,
                obj_id,
                "Yes" if text1 or text2 else "No",  # If either found a plate
                confidence1,
                f"{obj_id}/{image_file1}" if image_file1 else "",
                confidence2,
                f"{obj_id}/{image_file2}" if image_file2 else "",
                text1,
                text2
            ]
            sheet.append(row)

    workbook.save(output_file)
    send_email_with_attachment(sender_email, to_emails, cc_emails, password, subject, body, output_file)
    logger.info(f"OCR process completed and results saved to Excel for {date}")


def scheduled_job():
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    ocr_process = mp.Process(target=run_ocr_and_save_to_excel, args=(yesterday,))
    ocr_process.start()
    ocr_process.join()


def main():
    logger.info("Starting main process")
    ensure_dir(OUTPUT_DIR)

    frame_queue = mp.Queue()
    result_queue = mp.Queue()

    vehicle_process = mp.Process(target=vehicle_detection, args=(frame_queue, result_queue))
    plate_process = mp.Process(target=plate_detection, args=(frame_queue, result_queue))

    vehicle_process.start()
    plate_process.start()

    # Schedule the OCR job to run daily at 00:01 AM
    schedule.every().day.at("00:01").do(scheduled_job)

    # Run the scheduled jobs
    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == '__main__':
    mp.set_start_method('spawn')
    main()